"""F6: phiếu hủy hàng (hết hạn, hỏng vỡ, thất thoát).

Vấn đề mà nó sửa: trước F6, hàng hết hạn chỉ có một đường ra là xuất kho
(`adjust_stock` delta âm). Đường đó không ghi lý do và không chốt giá vốn ở đâu,
nên số hàng đó **biến mất khỏi báo cáo**: tồn giảm, doanh thu không đổi, và lãi
gộp cao hơn thực tế đúng bằng phần vốn vừa mất. Sai theo hướng làm người xem yên
tâm - cùng kiểu với bẫy 13 (giá vốn NULL bị đọc thành 0).

`test_huy_hang_lam_giam_lai_dung_bang_gia_von` là test quan trọng nhất ở đây.
"""
from datetime import datetime, timedelta
from io import BytesIO

import openpyxl
from conftest import _unique, admin_token, auth, new_seller, new_staff, seller_with_shop

from fselling import models
from fselling.core import thoi_gian
from fselling.core.database import SessionLocal
from fselling.services import inventory_service


def _ngay(cach_hom_nay: int) -> str:
    """Ngày cách hôm nay N ngày, tính theo NGÀY NGHIỆP VỤ Việt Nam (bẫy 36).

    Các ca ở đây đều lùi từ -1 trở đi nên bản cũ tính bằng `utcnow()` vẫn xanh,
    nhưng giữ nguyên là để sẵn một cái bẫy: ai thêm ca `_ngay(0)` sau này sẽ
    thấy test đỏ ngẫu nhiên trong khung 0h-7h sáng mà không hiểu vì sao.
    """
    return (thoi_gian.hom_nay_vn() + timedelta(days=cach_hom_nay)).isoformat()


def _sp(product_id):
    session = SessionLocal()
    try:
        return (
            session.query(models.Product)
            .filter(models.Product.id == product_id)
            .first()
        )
    finally:
        session.close()


def _lo(product_id):
    session = SessionLocal()
    try:
        return (
            session.query(models.ProductBatch)
            .filter(models.ProductBatch.product_id == product_id)
            .order_by(models.ProductBatch.id)
            .all()
        )
    finally:
        session.close()


def _khong_lech(shop_id):
    session = SessionLocal()
    try:
        return inventory_service.doi_chieu_ton_kho(session, shop_id) == []
    finally:
        session.close()


def _tao_sp_theo_lo(client, ctx, gia_ban=25000):
    res = client.post(
        "/api/products",
        params={"shop_id": ctx["shop_id"]},
        data={
            "name": _unique("Sua tuoi"),
            "price": gia_ban,
            "stock": 0,
            "category_id": ctx["category_id"],
            "track_batches": "true",
        },
        headers=auth(ctx["token"]),
    )
    assert res.status_code == 200, res.text
    return res.json()


def _nhap_lo(client, ctx, product_id, so_luong, han, gia_von=None):
    body = {
        "delta": so_luong,
        "expiry_date": han,
        "reason": "Kiểm thử nhập lô",
    }
    if gia_von is not None:
        body["unit_cost"] = gia_von
    res = client.post(
        f"/api/products/{product_id}/stock", json=body, headers=auth(ctx["token"])
    )
    assert res.status_code == 200, res.text
    return res.json()


def _huy(client, token, shop_id, items, reason="EXPIRED", **kwargs):
    body = {"reason": reason, "items": items}
    body.update(kwargs)
    return client.post(
        f"/api/products/{shop_id}/write-off", json=body, headers=auth(token)
    )


def _stats(client, ctx, token=None):
    res = client.get(
        f"/api/shops/{ctx['shop_id']}/stats",
        headers=auth(token or ctx["token"]),
    )
    assert res.status_code == 200, res.text
    return res.json()


# ---------- Hủy hàng theo lô ----------


def test_huy_lo_het_han_tru_dung_lo_do(client):
    ctx = seller_with_shop(client)
    sp = _tao_sp_theo_lo(client, ctx)
    _nhap_lo(client, ctx, sp["id"], 6, _ngay(-3), gia_von=10000)   # đã hỏng
    _nhap_lo(client, ctx, sp["id"], 10, _ngay(120), gia_von=12000)  # còn tốt
    lo = _lo(sp["id"])

    res = _huy(client, ctx["token"], ctx["shop_id"], [
        {"product_id": sp["id"], "batch_id": lo[0].id, "quantity": 6}
    ])
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["total_quantity"] == 6
    assert body["total_cost"] == 60000
    assert body["items"][0]["expiry_date"] == _ngay(-3)

    assert [b.quantity for b in _lo(sp["id"])] == [0, 10]
    assert _sp(sp["id"]).stock == 10
    assert _khong_lech(ctx["shop_id"])


def test_huy_mot_phan_cua_lo(client):
    ctx = seller_with_shop(client)
    sp = _tao_sp_theo_lo(client, ctx)
    _nhap_lo(client, ctx, sp["id"], 10, _ngay(-1), gia_von=8000)
    lo_id = _lo(sp["id"])[0].id

    res = _huy(client, ctx["token"], ctx["shop_id"], [
        {"product_id": sp["id"], "batch_id": lo_id, "quantity": 4}
    ], reason="DAMAGED")
    assert res.status_code == 200, res.text
    assert res.json()["total_cost"] == 32000

    assert [b.quantity for b in _lo(sp["id"])] == [6]
    assert _sp(sp["id"]).stock == 6
    assert _khong_lech(ctx["shop_id"])


def test_huy_qua_so_lo_dang_co_bi_tu_choi(client):
    ctx = seller_with_shop(client)
    sp = _tao_sp_theo_lo(client, ctx)
    _nhap_lo(client, ctx, sp["id"], 5, _ngay(-1))
    lo_id = _lo(sp["id"])[0].id

    res = _huy(client, ctx["token"], ctx["shop_id"], [
        {"product_id": sp["id"], "batch_id": lo_id, "quantity": 6}
    ])
    assert res.status_code == 400
    assert _sp(sp["id"]).stock == 5


def test_huy_hang_theo_lo_ma_khong_chon_lo_bi_tu_choi(client):
    """Không nói lô nào thì không biết chốt giá vốn nào, và cũng không biết hạn
    nào vừa bị bỏ đi."""
    ctx = seller_with_shop(client)
    sp = _tao_sp_theo_lo(client, ctx)
    _nhap_lo(client, ctx, sp["id"], 5, _ngay(-1))

    res = _huy(client, ctx["token"], ctx["shop_id"], [
        {"product_id": sp["id"], "quantity": 2}
    ])
    assert res.status_code == 400
    assert "chọn lô" in res.json()["detail"]
    assert _sp(sp["id"]).stock == 5


def test_huy_lo_cua_san_pham_khac_bi_tu_choi(client):
    ctx = seller_with_shop(client)
    a = _tao_sp_theo_lo(client, ctx)
    b = _tao_sp_theo_lo(client, ctx)
    _nhap_lo(client, ctx, a["id"], 5, _ngay(-1))
    _nhap_lo(client, ctx, b["id"], 5, _ngay(-1))
    lo_cua_b = _lo(b["id"])[0].id

    res = _huy(client, ctx["token"], ctx["shop_id"], [
        {"product_id": a["id"], "batch_id": lo_cua_b, "quantity": 1}
    ])
    assert res.status_code == 404
    assert _sp(b["id"]).stock == 5


# ---------- Hủy hàng không theo lô ----------


def test_huy_hang_khong_theo_lo(client):
    ctx = seller_with_shop(client)
    sp = ctx["product"]
    client.put(
        f"/api/products/{sp['id']}",
        data={
            "name": sp["name"],
            "price": sp["price"],
            "category_id": ctx["category_id"],
            "cost_price": 40000,
        },
        headers=auth(ctx["token"]),
    )

    res = _huy(client, ctx["token"], ctx["shop_id"], [
        {"product_id": sp["id"], "quantity": 3}
    ], reason="LOST")
    assert res.status_code == 200, res.text
    assert res.json()["total_cost"] == 120000
    assert _sp(sp["id"]).stock == 7


def test_huy_hang_khong_theo_lo_ma_chon_lo_bi_tu_choi(client):
    ctx = seller_with_shop(client)
    res = _huy(client, ctx["token"], ctx["shop_id"], [
        {"product_id": ctx["product"]["id"], "batch_id": 1, "quantity": 1}
    ])
    assert res.status_code == 400
    assert _sp(ctx["product"]["id"]).stock == 10


def test_huy_qua_ton_kho_bi_tu_choi(client):
    ctx = seller_with_shop(client)
    res = _huy(client, ctx["token"], ctx["shop_id"], [
        {"product_id": ctx["product"]["id"], "quantity": 11}
    ])
    assert res.status_code == 400
    assert _sp(ctx["product"]["id"]).stock == 10


# ---------- Kiểm dữ liệu vào ----------


def test_ly_do_la_bi_tu_choi(client):
    ctx = seller_with_shop(client)
    res = _huy(client, ctx["token"], ctx["shop_id"], [
        {"product_id": ctx["product"]["id"], "quantity": 1}
    ], reason="TUY_Y")
    assert res.status_code == 400
    assert _sp(ctx["product"]["id"]).stock == 10


def test_so_luong_khong_duong_bi_tu_choi(client):
    ctx = seller_with_shop(client)
    for so in (0, -2):
        res = _huy(client, ctx["token"], ctx["shop_id"], [
            {"product_id": ctx["product"]["id"], "quantity": so}
        ])
        assert res.status_code == 400
    assert _sp(ctx["product"]["id"]).stock == 10


def test_phieu_rong_bi_tu_choi(client):
    ctx = seller_with_shop(client)
    assert _huy(client, ctx["token"], ctx["shop_id"], []).status_code == 400


def test_mot_lo_hai_dong_trong_cung_phieu_bi_tu_choi(client):
    """Phép kiểm 'đủ hàng để hủy' xét từng dòng riêng lẻ, nên hai dòng cùng lô
    có thể cùng lọt trong khi tổng vượt quá số đang có."""
    ctx = seller_with_shop(client)
    sp = _tao_sp_theo_lo(client, ctx)
    _nhap_lo(client, ctx, sp["id"], 5, _ngay(-1))
    lo_id = _lo(sp["id"])[0].id

    res = _huy(client, ctx["token"], ctx["shop_id"], [
        {"product_id": sp["id"], "batch_id": lo_id, "quantity": 4},
        {"product_id": sp["id"], "batch_id": lo_id, "quantity": 4},
    ])
    assert res.status_code == 400
    assert _sp(sp["id"]).stock == 5


def test_khong_huy_duoc_hang_cua_shop_khac(client):
    ctx = seller_with_shop(client)
    khac = seller_with_shop(client)

    res = _huy(client, ctx["token"], ctx["shop_id"], [
        {"product_id": khac["product"]["id"], "quantity": 1}
    ])
    assert res.status_code == 404
    assert _sp(khac["product"]["id"]).stock == 10


# ---------- Chống bấm hai lần ----------


def test_gui_lai_cung_ma_thao_tac_khong_tru_kho_hai_lan(client):
    ctx = seller_with_shop(client)
    dong = [{"product_id": ctx["product"]["id"], "quantity": 3}]

    dau = _huy(client, ctx["token"], ctx["shop_id"], dong, operation_id="huy-1")
    assert dau.status_code == 200, dau.text
    assert dau.json()["repeated"] is False
    assert _sp(ctx["product"]["id"]).stock == 7

    lai = _huy(client, ctx["token"], ctx["shop_id"], dong, operation_id="huy-1")
    assert lai.status_code == 200, lai.text
    assert lai.json()["repeated"] is True
    assert lai.json()["write_off_id"] == dau.json()["write_off_id"]
    assert _sp(ctx["product"]["id"]).stock == 7, "KHÔNG được trừ lần hai"


def test_ma_thao_tac_dung_lai_o_shop_khac_bi_tu_choi(client):
    ctx = seller_with_shop(client)
    khac = seller_with_shop(client)

    _huy(client, ctx["token"], ctx["shop_id"],
         [{"product_id": ctx["product"]["id"], "quantity": 1}], operation_id="chung")
    res = _huy(client, khac["token"], khac["shop_id"],
               [{"product_id": khac["product"]["id"], "quantity": 1}],
               operation_id="chung")
    assert res.status_code == 409
    assert _sp(khac["product"]["id"]).stock == 10


# ---------- Phân quyền ----------


def test_nhan_vien_khong_huy_duoc_hang(client):
    """Hủy hàng là đường duy nhất làm tồn giảm mà không sinh doanh thu, nên nó
    cũng là đường thuận tiện nhất để che hàng thất thoát."""
    ctx = seller_with_shop(client)
    for vai in ("MANAGER", "WAREHOUSE", "CASHIER"):
        _, token = new_staff(client, ctx, staff_role=vai)
        res = _huy(client, token, ctx["shop_id"], [
            {"product_id": ctx["product"]["id"], "quantity": 1}
        ])
        assert res.status_code == 403, f"{vai} không được hủy hàng"
    assert _sp(ctx["product"]["id"]).stock == 10


def test_admin_huy_duoc_hang(client):
    ctx = seller_with_shop(client)
    res = _huy(client, admin_token(client), ctx["shop_id"], [
        {"product_id": ctx["product"]["id"], "quantity": 2}
    ])
    assert res.status_code == 200, res.text
    assert _sp(ctx["product"]["id"]).stock == 8


def test_khong_dang_nhap_thi_khong_huy_duoc(client):
    ctx = seller_with_shop(client)
    res = client.post(
        f"/api/products/{ctx['shop_id']}/write-off",
        json={"reason": "LOST", "items": [
            {"product_id": ctx["product"]["id"], "quantity": 1}
        ]},
    )
    assert res.status_code == 401


# ---------- Đề xuất hủy hàng hết hạn ----------


def test_de_xuat_chi_lay_lo_da_qua_han(client):
    ctx = seller_with_shop(client)
    sp = _tao_sp_theo_lo(client, ctx)
    _nhap_lo(client, ctx, sp["id"], 3, _ngay(-10), gia_von=10000)
    _nhap_lo(client, ctx, sp["id"], 4, _ngay(2), gia_von=10000)     # sắp hết
    _nhap_lo(client, ctx, sp["id"], 5, _ngay(300), gia_von=10000)

    res = client.get(
        f"/api/products/{ctx['shop_id']}/write-off/expired",
        headers=auth(ctx["token"]),
    )
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["total_quantity"] == 3
    assert body["total_cost"] == 30000
    assert [d["expiry_date"] for d in body["items"]] == [_ngay(-10)]


def test_de_xuat_khong_tu_huy(client):
    """Hạn nhập sai một chữ số là cả lô còn tốt bị bỏ đi, mà hủy không có đường
    lùi. Đề xuất phải đi qua mắt người."""
    ctx = seller_with_shop(client)
    sp = _tao_sp_theo_lo(client, ctx)
    _nhap_lo(client, ctx, sp["id"], 7, _ngay(-4), gia_von=5000)

    client.get(
        f"/api/products/{ctx['shop_id']}/write-off/expired",
        headers=auth(ctx["token"]),
    )
    assert _sp(sp["id"]).stock == 7


def test_de_xuat_het_han_het_thi_tra_ve_rong(client):
    ctx = seller_with_shop(client)
    res = client.get(
        f"/api/products/{ctx['shop_id']}/write-off/expired",
        headers=auth(ctx["token"]),
    )
    assert res.status_code == 200, res.text
    assert res.json() == {"items": [], "total_quantity": 0, "total_cost": 0}


def test_nhan_vien_khong_xem_duoc_de_xuat_va_danh_sach_phieu(client):
    ctx = seller_with_shop(client)
    _, manager = new_staff(client, ctx, staff_role="MANAGER")
    for duong in ("write-off/expired", "write-offs"):
        res = client.get(
            f"/api/products/{ctx['shop_id']}/{duong}", headers=auth(manager)
        )
        assert res.status_code == 403, duong


def test_phieu_ghi_lai_AI_bam_huy(client):
    """Màn kiểm toán mà không biết ai bấm thì không kiểm được gì.

    Hủy hàng là đường duy nhất làm tồn giảm mà không sinh doanh thu, nên "ai
    làm" quan trọng ngang "mất bao nhiêu".
    """
    ctx = seller_with_shop(client)
    res = _huy(client, ctx["token"], ctx["shop_id"], [
        {"product_id": ctx["product"]["id"], "quantity": 1}
    ])
    assert res.status_code == 200, res.text
    assert res.json()["created_by"] == ctx["username"]

    ds = client.get(
        f"/api/products/{ctx['shop_id']}/write-offs", headers=auth(ctx["token"])
    ).json()["write_offs"]
    assert ds[0]["created_by"] == ctx["username"]


def test_admin_huy_thi_ghi_ten_admin(client):
    ctx = seller_with_shop(client)
    _huy(client, admin_token(client), ctx["shop_id"], [
        {"product_id": ctx["product"]["id"], "quantity": 1}
    ])
    ds = client.get(
        f"/api/products/{ctx['shop_id']}/write-offs", headers=auth(ctx["token"])
    ).json()["write_offs"]
    assert ds[0]["created_by"] == "admin"


def test_danh_sach_phieu_moi_nhat_truoc(client):
    ctx = seller_with_shop(client)
    _huy(client, ctx["token"], ctx["shop_id"],
         [{"product_id": ctx["product"]["id"], "quantity": 1}], reason="LOST")
    _huy(client, ctx["token"], ctx["shop_id"],
         [{"product_id": ctx["product"]["id"], "quantity": 2}], reason="DAMAGED")

    res = client.get(
        f"/api/products/{ctx['shop_id']}/write-offs", headers=auth(ctx["token"])
    )
    assert res.status_code == 200, res.text
    phieu = res.json()["write_offs"]
    assert [p["reason"] for p in phieu] == ["DAMAGED", "LOST"]
    assert [p["total_quantity"] for p in phieu] == [2, 1]


# ---------- File Excel ----------
#
# Sheet đơn hàng đọc một mình thì tồn kho giảm mà không hiểu vì sao, và số lãi
# gộp ở cuối sheet đó cao hơn con số trên Dashboard đúng bằng phần vốn đã hủy.
# Hai chỗ nói hai số khác nhau về cùng một tháng là lúc người ta thôi tin cả hai.


def _excel(client, ctx, token=None, lang=None):
    headers = auth(token or ctx["token"])
    if lang:
        headers["Accept-Language"] = lang
    res = client.get(f"/api/export/seller/{ctx['shop_id']}", headers=headers)
    assert res.status_code == 200, res.text
    return openpyxl.load_workbook(BytesIO(res.content))


def test_excel_co_sheet_hang_da_huy(client):
    ctx = seller_with_shop(client)
    sp = _tao_sp_theo_lo(client, ctx)
    _nhap_lo(client, ctx, sp["id"], 6, _ngay(-3), gia_von=10000)
    lo_id = _lo(sp["id"])[0].id
    _huy(client, ctx["token"], ctx["shop_id"], [
        {"product_id": sp["id"], "batch_id": lo_id, "quantity": 6}
    ], note="Sua qua han cuoi thang")

    wb = _excel(client, ctx)
    assert "Hàng đã hủy" in wb.sheetnames
    ws = wb["Hàng đã hủy"]
    tieu_de = [c.value for c in ws[1]]
    assert tieu_de[:6] == [
        "Mã phiếu", "Ngày hủy", "Lý do", "Sản phẩm", "Hạn sử dụng", "Số lượng"
    ]

    dong = [c.value for c in ws[2]]
    assert dong[2] == "Hết hạn"
    assert dong[4] == _ngay(-3)
    assert dong[5] == 6
    assert dong[6] == 10000          # giá vốn đơn vị
    assert dong[7] == 60000          # thành tiền
    assert dong[8] == ctx["username"]
    assert dong[9] == "Sua qua han cuoi thang"


def test_excel_mot_dong_cho_MOT_LO(client):
    """Đối chiếu cuối tháng cần biết từng hạn, không phải một ô gộp."""
    ctx = seller_with_shop(client)
    sp = _tao_sp_theo_lo(client, ctx)
    _nhap_lo(client, ctx, sp["id"], 4, _ngay(-5), gia_von=1000)
    _nhap_lo(client, ctx, sp["id"], 3, _ngay(-2), gia_von=2000)
    lo = _lo(sp["id"])
    _huy(client, ctx["token"], ctx["shop_id"], [
        {"product_id": sp["id"], "batch_id": lo[0].id, "quantity": 4},
        {"product_id": sp["id"], "batch_id": lo[1].id, "quantity": 3},
    ])

    ws = _excel(client, ctx)["Hàng đã hủy"]
    han = [ws.cell(row=r, column=5).value for r in (2, 3)]
    assert sorted(han) == sorted([_ngay(-5), _ngay(-2)])


def test_excel_phieu_thieu_gia_von_de_TRONG_va_dem_rieng(client):
    """Ô trống chứ không phải 0: 0 trong cột tiền đọc ra là "không đáng đồng
    nào", khác hẳn "chưa ai khai"."""
    ctx = seller_with_shop(client)
    sp = _tao_sp_theo_lo(client, ctx)
    _nhap_lo(client, ctx, sp["id"], 5, _ngay(-2))     # KHÔNG khai giá vốn
    lo_id = _lo(sp["id"])[0].id
    _huy(client, ctx["token"], ctx["shop_id"], [
        {"product_id": sp["id"], "batch_id": lo_id, "quantity": 5}
    ])

    ws = _excel(client, ctx)["Hàng đã hủy"]
    # openpyxl đọc ô rỗng ra `None`. Điều quan trọng là nó KHÔNG phải số 0.
    assert ws.cell(row=2, column=7).value is None
    assert ws.cell(row=2, column=8).value == "Chưa khai giá vốn"

    cot_a = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    assert "Số phiếu chưa đủ giá vốn (không tính vào lỗ)" in cot_a
    i = cot_a.index("Tổng lỗ (phiếu đã đủ giá vốn)")
    assert ws.cell(row=i + 1, column=2).value == 0, "Không đoán giá vốn"


def test_excel_cua_nhan_vien_KHONG_co_sheet_hang_huy(client):
    """Số lỗ chính là giá vốn nhân số lượng, nói ra nó là nói ra giá vốn."""
    ctx = seller_with_shop(client)
    _huy(client, ctx["token"], ctx["shop_id"], [
        {"product_id": ctx["product"]["id"], "quantity": 1}
    ])
    _, manager = new_staff(client, ctx, staff_role="MANAGER")

    wb = _excel(client, ctx, token=manager)
    assert "Hàng đã hủy" not in wb.sheetnames


def test_excel_tieng_anh_dich_ca_sheet_hang_huy(client):
    ctx = seller_with_shop(client)
    _huy(client, ctx["token"], ctx["shop_id"], [
        {"product_id": ctx["product"]["id"], "quantity": 2}
    ], reason="DAMAGED")

    wb = _excel(client, ctx, lang="en")
    assert "Written off" in wb.sheetnames
    ws = wb["Written off"]
    assert [c.value for c in ws[1]][:6] == [
        "Write-off ID", "Date", "Reason", "Product", "Expiry date", "Quantity"
    ]
    assert ws.cell(row=2, column=3).value == "Damaged"


def test_excel_khong_co_phieu_huy_thi_sheet_van_co_tieu_de(client):
    ctx = seller_with_shop(client)
    ws = _excel(client, ctx)["Hàng đã hủy"]
    assert [c.value for c in ws[1]][0] == "Mã phiếu"
    assert ws.max_row == 1


# ---------- Ảnh hưởng tới báo cáo ----------


def test_huy_hang_lam_giam_lai_dung_bang_gia_von(client):
    """Đây là lý do tồn tại của cả tính năng.

    Bán 1 món giá 100.000 giá vốn 60.000 -> lãi 40.000. Hủy tiếp 2 món cùng giá
    vốn thì lỗ thêm 120.000, lãi còn -80.000. Trước F6 số hàng hủy đi qua đường
    xuất kho và lãi vẫn báo 40.000.
    """
    ctx = seller_with_shop(client)
    sp = ctx["product"]
    client.put(
        f"/api/products/{sp['id']}",
        data={
            "name": sp["name"],
            "price": sp["price"],
            "category_id": ctx["category_id"],
            "cost_price": 60000,
        },
        headers=auth(ctx["token"]),
    )
    don = client.post(
        f"/api/orders/{ctx['shop_id']}",
        json={
            "items": [{"product_id": sp["id"], "price": 100000, "quantity": 1}],
            "payment_method": "cash",
        },
        headers=auth(ctx["token"]),
    ).json()
    client.post(f"/api/orders/{don['order_id']}/pay", headers=auth(ctx["token"]))

    truoc = _stats(client, ctx)
    assert truoc["gross_profit"] == 40000
    assert truoc["write_off_loss"] == 0
    assert truoc["written_off_quantity"] == 0

    res = _huy(client, ctx["token"], ctx["shop_id"], [
        {"product_id": sp["id"], "quantity": 2}
    ])
    assert res.status_code == 200, res.text

    sau = _stats(client, ctx)
    assert sau["write_off_loss"] == 120000
    assert sau["written_off_quantity"] == 2
    assert sau["gross_profit"] == 40000 - 120000
    assert sau["total_revenue"] == 100000, "Hủy hàng KHÔNG đụng vào doanh thu"


def test_phieu_thieu_gia_von_bi_loai_va_dem_rieng(client):
    """NULL không bao giờ được đọc thành 0. Cộng phần biết được rồi trình bày
    như tổng thiệt hại là báo lỗ THẤP hơn thực tế."""
    ctx = seller_with_shop(client)
    sp = _tao_sp_theo_lo(client, ctx)
    _nhap_lo(client, ctx, sp["id"], 5, _ngay(-2))       # KHÔNG khai giá vốn
    lo_id = _lo(sp["id"])[0].id

    res = _huy(client, ctx["token"], ctx["shop_id"], [
        {"product_id": sp["id"], "batch_id": lo_id, "quantity": 5}
    ])
    assert res.status_code == 200, res.text
    assert res.json()["total_cost"] is None

    stats = _stats(client, ctx)
    assert stats["write_off_loss"] == 0, "Không đoán giá vốn"
    assert stats["write_offs_missing_cost"] == 1
    assert stats["written_off_quantity"] == 5, "Số lượng thì luôn biết chắc"


def test_nhan_vien_khong_thay_so_lo_huy_hang(client):
    """Số lỗ chính là giá vốn nhân số lượng, nói ra nó là nói ra giá vốn."""
    ctx = seller_with_shop(client)
    _huy(client, ctx["token"], ctx["shop_id"],
         [{"product_id": ctx["product"]["id"], "quantity": 1}])
    _, manager = new_staff(client, ctx, staff_role="MANAGER")

    body = _stats(client, ctx, token=manager)
    assert "total_revenue" in body
    for khoa in ("write_off_loss", "written_off_quantity", "write_offs_missing_cost"):
        assert khoa not in body, khoa


def test_gia_von_chot_theo_dung_lo_bi_huy(client):
    """Lô nhập đắt hỏng trên kệ là mất đúng số tiền của lô đó, không phải bình
    quân của sản phẩm."""
    ctx = seller_with_shop(client)
    sp = _tao_sp_theo_lo(client, ctx)
    _nhap_lo(client, ctx, sp["id"], 10, _ngay(-1), gia_von=30000)   # lô đắt, hỏng
    _nhap_lo(client, ctx, sp["id"], 90, _ngay(300), gia_von=1000)   # lô rẻ, còn tốt
    lo_dat = _lo(sp["id"])[0].id

    res = _huy(client, ctx["token"], ctx["shop_id"], [
        {"product_id": sp["id"], "batch_id": lo_dat, "quantity": 10}
    ])
    assert res.status_code == 200, res.text
    assert res.json()["total_cost"] == 300000
    assert _stats(client, ctx)["write_off_loss"] == 300000
