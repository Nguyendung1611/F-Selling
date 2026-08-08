"""Dashboard, thống kê và xuất Excel."""
from __future__ import annotations

import io
from datetime import date, datetime, timedelta, timezone
from typing import Any, Dict, List, Optional
from zoneinfo import ZoneInfo

import openpyxl
from fastapi import HTTPException
from sqlalchemy import and_, distinct, func, or_
from sqlalchemy.orm import Session, joinedload

from .. import models
from ..core import thoi_gian
from ..core.config import log_to_file
from ..core.i18n import tr
from ..dependencies import (
    PERMISSION_REPORT,
    has_cost_visibility,
    require_cost_visibility,
    require_shop_access,
    require_staff_permission,
)
from . import (
    expense_service,
    order_service,
    return_service,
    shift_service,
    subscription_service,
    write_off_service,
)

TREND_DAYS = 7
FREE_REPORT_DAYS = 31
_VIETNAM_TZ = ZoneInfo("Asia/Ho_Chi_Minh")


def _today_vietnam():
    """Ngày nghiệp vụ theo múi giờ mà giao diện và cửa hàng đang sử dụng."""
    return thoi_gian.hom_nay_vn()


def _paid_revenue(db: Session, shop_id: int) -> float:
    return (
        db.query(func.sum(models.Order.total_amount))
        .filter(models.Order.shop_id == shop_id, models.Order.status == "PAID")
        .scalar()
        or 0
    )


DEFAULT_PAGE_SIZE = 50
MAX_PAGE_SIZE = 200


def _parse_ngay(chuoi: Optional[str], ten_truong: str) -> Optional[date]:
    """Chuỗi YYYY-MM-DD -> ngày lịch. Sai định dạng -> 400 rõ ràng."""
    if not chuoi or not chuoi.strip():
        return None
    try:
        return datetime.strptime(chuoi.strip(), "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(
            status_code=400,
            detail=tr(
                "{field} phải theo định dạng YYYY-MM-DD",
                field=ten_truong,
            ),
        )


def _dau_ngay_viet_nam_sang_utc(ngay: date) -> datetime:
    """00:00 ngày Việt Nam -> UTC-naive, cùng chuẩn lưu ``created_at``."""
    return thoi_gian.dau_ngay_vn_sang_utc(ngay)


def _loc_khoang_ngay(
    query, tu_ngay: Optional[str], den_ngay: Optional[str], cot=None
):
    """Lọc theo created_at. `den_ngay` tính trọn cả ngày đó (đến 23:59:59).

    `cot` cho phép lọc bảng khác: phiếu trả hàng phải tính theo NGÀY TRẢ, không
    phải ngày bán - nếu không thì một lần trả hôm nay sẽ đi sửa ngược con số
    lãi của tháng trước, thứ đã chốt sổ rồi.
    """
    if cot is None:
        cot = models.Order.created_at
    bat_dau = _parse_ngay(tu_ngay, "tu_ngay")
    ket_thuc = _parse_ngay(den_ngay, "den_ngay")
    if bat_dau and ket_thuc and bat_dau > ket_thuc:
        raise HTTPException(
            status_code=400,
            detail=tr("tu_ngay không được lớn hơn den_ngay"),
        )
    if bat_dau:
        query = query.filter(cot >= _dau_ngay_viet_nam_sang_utc(bat_dau))
    if ket_thuc:
        query = query.filter(
            cot < _dau_ngay_viet_nam_sang_utc(ket_thuc + timedelta(days=1))
        )
    return query


def _limit_free_report_range(
    db: Session,
    shop_id: int,
    tu_ngay: Optional[str],
    den_ngay: Optional[str],
) -> tuple[Optional[str], Optional[str]]:
    """Free xem 31 ngày gần nhất; Pro giữ nguyên khoảng người dùng chọn.

    Nếu người dùng bấm chọn một kỳ cũ thì nói rõ cần Pro, không âm thầm đổi kỳ
    rồi đưa ra con số khác với thứ họ vừa yêu cầu. Khi không chọn ngày, tự đặt
    mốc đầu 31 ngày để dashboard Free không vô tình đọc toàn bộ lịch sử.
    """
    if subscription_service.get_subscription_state(db, shop_id)["can_use_pro"]:
        return tu_ngay, den_ngay

    cutoff = _today_vietnam() - timedelta(days=FREE_REPORT_DAYS - 1)
    parsed_start = _parse_ngay(tu_ngay, "tu_ngay")
    parsed_end = _parse_ngay(den_ngay, "den_ngay")
    if (parsed_start is not None and parsed_start < cutoff) or (
        parsed_end is not None and parsed_end < cutoff
    ):
        raise HTTPException(
            status_code=402,
            detail=tr(
                "Gói Free xem báo cáo trong 31 ngày gần nhất. Hãy mở tab Gói cước để xem kỳ cũ hơn."
            ),
        )
    return tu_ngay or cutoff.isoformat(), den_ngay


def _cong_no_phai_thu(db: Session, shop_id: int) -> float:
    """Tổng tiền khách còn nợ shop, tại thời điểm hiện tại."""
    rows = (
        db.query(
            models.Order.total_amount,
            models.Order.paid_amount,
            models.Order.cash_paid_amount,
        )
        .filter(
            models.Order.shop_id == shop_id,
            models.Order.status == order_service.STATUS_DEBT,
        )
        .all()
    )
    return sum(
        max(float(tong or 0) - float(bank or 0) - float(tien_mat or 0), 0.0)
        for tong, bank, tien_mat in rows
    )


def _phieu_tra_trong_ky(
    db: Session, shop_id: int, tu_ngay: Optional[str], den_ngay: Optional[str]
):
    """Query các phiếu trả hàng của shop theo NGÀY TRẢ."""
    return _loc_khoang_ngay(
        db.query(models.OrderReturn).filter(
            models.OrderReturn.shop_id == shop_id
        ),
        tu_ngay,
        den_ngay,
        cot=models.OrderReturn.created_at,
    )


def _huy_hang_anh_huong_lai(
    db: Session, shop_id: int, tu_ngay: Optional[str], den_ngay: Optional[str]
) -> Dict[str, Any]:
    """Khoản lỗ do hủy hàng (hết hạn, hỏng vỡ, thất thoát), theo NGÀY HỦY.

    Lỗ đúng bằng giá vốn số hàng đã bỏ đi - hàng hủy không sinh đồng doanh thu
    nào nên không có gì để trừ ra như ở trả hàng.

    Đây là chỗ mà trước F6 báo cáo nói dối: hàng hết hạn đi qua đường xuất kho
    thì tồn giảm, doanh thu không đổi, và lãi gộp cao hơn thực tế đúng bằng phần
    vốn vừa mất. Con số sai theo hướng làm người xem yên tâm nên rất khó nghi.

    Phiếu nào còn dòng chưa khai giá vốn thì bị loại NGUYÊN PHIẾU và đếm riêng -
    cùng nguyên tắc "không đoán NULL là 0" ở `_lai_gop` và `_tra_hang_anh_huong_lai`.
    Trừ phần biết được rồi trình bày như tổng thiệt hại là báo lỗ THẤP hơn thực
    tế, tức là vẫn sai theo đúng hướng đó.
    """
    phieu = _loc_khoang_ngay(
        db.query(models.StockWriteOff).filter(
            models.StockWriteOff.shop_id == shop_id
        ),
        tu_ngay,
        den_ngay,
        cot=models.StockWriteOff.created_at,
    ).all()
    if not phieu:
        return {
            "written_off_quantity": 0,
            "write_off_loss": 0.0,
            "write_offs_missing_cost": 0,
        }

    ids = [p.id for p in phieu]
    dong_theo_phieu: Dict[int, List[models.StockWriteOffItem]] = {}
    for d in (
        db.query(models.StockWriteOffItem)
        .filter(models.StockWriteOffItem.write_off_id.in_(ids))
        .all()
    ):
        dong_theo_phieu.setdefault(d.write_off_id, []).append(d)

    lo = 0.0
    thieu_gia_von = 0
    for p in phieu:
        dong = dong_theo_phieu.get(p.id, [])
        if any(d.cost_price is None for d in dong):
            thieu_gia_von += 1
            continue
        lo += sum(float(d.cost_price) * int(d.quantity or 0) for d in dong)

    return {
        # Số lượng thì đếm ĐỦ mọi phiếu, kể cả phiếu thiếu giá vốn: "đã bỏ đi
        # bao nhiêu món" luôn biết chắc, chỉ "mất bao nhiêu tiền" mới cần giá vốn.
        "written_off_quantity": sum(int(p.total_quantity or 0) for p in phieu),
        "write_off_loss": lo,
        "write_offs_missing_cost": thieu_gia_von,
    }


def _tra_hang_anh_huong_lai(
    db: Session, shop_id: int, tu_ngay: Optional[str], den_ngay: Optional[str]
) -> Dict[str, Any]:
    """Phần lãi bị mất vì hàng trả lại, tính theo NGÀY TRẢ.

    Lãi giảm đúng bằng: tiền đã hoàn - giá vốn thu hồi được.

    "Thu hồi được" chỉ tính những dòng ĐÃ NHẬP LẠI KHO. Hàng hỏng, bẩn, hết hạn
    không quay lại kệ nghĩa là shop mất trắng cả tiền hoàn lẫn vốn của món đó -
    lãi phải giảm bằng toàn bộ tiền hoàn. Tính gộp cả hai kiểu là báo lãi cao
    hơn thực tế đúng bằng giá vốn số hàng đã bỏ đi.

    Phiếu nào còn dòng nhập lại kho mà không biết giá vốn thì bị loại khỏi phần
    điều chỉnh và đếm riêng - cùng nguyên tắc "không đoán NULL là 0" ở `_lai_gop`.
    """
    phieu = _phieu_tra_trong_ky(db, shop_id, tu_ngay, den_ngay).all()
    tong_hoan = sum(float(p.refund_amount or 0) for p in phieu)
    if not phieu:
        return {
            "returned_amount": 0.0,
            "profit_reduction": 0.0,
            "returns_missing_cost": 0,
        }

    ids = [p.id for p in phieu]
    dong_theo_phieu: Dict[int, List[models.OrderReturnItem]] = {}
    for d in (
        db.query(models.OrderReturnItem)
        .filter(models.OrderReturnItem.return_id.in_(ids))
        .all()
    ):
        dong_theo_phieu.setdefault(d.return_id, []).append(d)

    giam_lai = 0.0
    thieu_gia_von = 0
    for p in phieu:
        dong = dong_theo_phieu.get(p.id, [])
        # Chỉ dòng nhập lại kho mới cần biết giá vốn; dòng bỏ đi thì mất trắng,
        # không phải tra giá vốn làm gì.
        if any(d.restocked and d.cost_price is None for d in dong):
            thieu_gia_von += 1
            continue
        von_thu_hoi = sum(
            float(d.cost_price or 0) * int(d.quantity or 0)
            for d in dong
            if d.restocked
        )
        giam_lai += float(p.refund_amount or 0) - von_thu_hoi

    return {
        "returned_amount": tong_hoan,
        "profit_reduction": giam_lai,
        "returns_missing_cost": thieu_gia_von,
    }


def _lai_gop(db: Session, paid_orders_subquery) -> Dict[str, Any]:
    """Lãi gộp của các đơn ĐÃ THANH TOÁN trong phạm vi truy vấn.

    Lãi gộp = doanh thu (đã trừ giảm giá voucher) - tổng giá vốn hàng bán.
    Giảm giá trừ ở mức ĐƠN HÀNG chứ không phân bổ xuống từng dòng, nên tổng số
    luôn khớp; đổi lại "lãi theo từng sản phẩm" (nếu sau này làm) sẽ là lãi chưa
    trừ giảm giá và phải nói rõ điều đó trên giao diện.

    Chỉ tính trên những đơn có ĐỦ giá vốn ở MỌI dòng. Đơn thiếu dù chỉ một dòng
    cũng bị loại nguyên đơn, vì giảm giá nằm ở mức đơn nên không tách được phần
    doanh thu tương ứng với riêng các dòng đã biết giá vốn. Loại nửa vời - trừ
    giá vốn đã biết ra khỏi toàn bộ doanh thu - còn tệ hơn không tính: nó ĐẨY
    LÃI LÊN đúng bằng phần chưa khai, và sai theo hướng làm người ta yên tâm.

    `cost_price` NULL không bao giờ được coi là 0. Phần bị loại trả về nguyên
    con số (`orders_missing_cost`, `revenue_missing_cost`) để giao diện nói ra
    báo cáo đang thiếu bao nhiêu, thay vì im lặng.
    """
    don_thieu_gia_von = db.query(distinct(models.OrderItem.order_id)).filter(
        models.OrderItem.order_id.in_(paid_orders_subquery),
        models.OrderItem.cost_price.is_(None),
    )

    def _tinh_duoc(query):
        """Giới hạn về các đơn đủ giá vốn."""
        return query.filter(~models.Order.id.in_(don_thieu_gia_von))

    doanh_thu_tinh_duoc = (
        _tinh_duoc(
            db.query(func.sum(models.Order.total_amount)).filter(
                models.Order.id.in_(paid_orders_subquery)
            )
        ).scalar()
        or 0
    )
    tong_gia_von = (
        db.query(func.sum(models.OrderItem.cost_price * models.OrderItem.quantity))
        .filter(
            models.OrderItem.order_id.in_(paid_orders_subquery),
            ~models.OrderItem.order_id.in_(don_thieu_gia_von),
        )
        .scalar()
        or 0
    )
    so_don_thieu = (
        db.query(func.count(models.Order.id))
        .filter(
            models.Order.id.in_(paid_orders_subquery),
            models.Order.id.in_(don_thieu_gia_von),
        )
        .scalar()
        or 0
    )
    doanh_thu_bi_loai = (
        db.query(func.sum(models.Order.total_amount))
        .filter(
            models.Order.id.in_(paid_orders_subquery),
            models.Order.id.in_(don_thieu_gia_von),
        )
        .scalar()
        or 0
    )

    lai = doanh_thu_tinh_duoc - tong_gia_von
    return {
        "revenue_with_cost": doanh_thu_tinh_duoc,
        "total_cost": tong_gia_von,
        "gross_profit": lai,
        # Doanh thu 0 thì tỷ suất không xác định, không phải 0%. Trả None để
        # giao diện hiện "--" thay vì một con số bịa.
        "gross_margin": (
            (lai / doanh_thu_tinh_duoc * 100) if doanh_thu_tinh_duoc else None
        ),
        "orders_missing_cost": so_don_thieu,
        "revenue_missing_cost": doanh_thu_bi_loai,
    }


def _don_co_tien_ve_chua_ghi_nhan(db: Session, shop_id: int):
    """Query id các đơn có tiền về (`BANK_UNAPPLIED`) mà người bán CHƯA xử lý.

    "Chưa xử lý" = kể từ bút toán tiền-về gần nhất, chưa có lần thu nợ nào được
    ghi. Lấy mốc theo thời gian chứ không theo số tiền: khách chuyển 40k của
    khoản nợ 100k rồi người bán ghi nhận 40k thì đơn vẫn còn nợ 60k, nhưng
    khoản chuyển kia đã được xử lý xong và không việc gì phải nhắc mãi.

    Trả về `Query` để caller đưa thẳng vào `in_()`; KHÔNG gọi `.subquery()`
    (bẫy 12 - SQLAlchemy 2.0 tự coerce, gọi tay sinh SAWarning).
    """
    tien_ve = (
        db.query(
            models.OrderPayment.order_id.label("order_id"),
            func.max(models.OrderPayment.created_at).label("moc"),
        )
        .filter(
            models.OrderPayment.entry_type == order_service.ENTRY_BANK_UNAPPLIED
        )
        .group_by(models.OrderPayment.order_id)
        .subquery()
    )
    da_xu_ly = (
        db.query(models.OrderPayment.order_id)
        .filter(
            models.OrderPayment.entry_type.in_(
                (
                    order_service.ENTRY_DEBT_CASH,
                    order_service.ENTRY_DEBT_TRANSFER,
                )
            ),
            models.OrderPayment.order_id == tien_ve.c.order_id,
            models.OrderPayment.created_at >= tien_ve.c.moc,
        )
    )
    return (
        db.query(models.Order.id)
        .join(tien_ve, tien_ve.c.order_id == models.Order.id)
        .filter(
            models.Order.shop_id == shop_id,
            ~models.Order.id.in_(da_xu_ly),
        )
    )


def seller_dashboard(
    db: Session,
    current_user: models.User,
    shop_id: int,
    page: int = 1,
    per_page: int = DEFAULT_PAGE_SIZE,
    tu_ngay: Optional[str] = None,
    den_ngay: Optional[str] = None,
    reconciliation_only: bool = False,
) -> Dict[str, Any]:
    """Danh sách đơn của shop, phân trang và lọc theo khoảng ngày.

    Không truyền tham số nào -> trang 1, 50 đơn mới nhất (như cũ với shop nhỏ).
    `total_revenue` luôn là doanh thu của KHOẢNG ĐANG LỌC, không phải của trang.
    """
    require_shop_access(db, shop_id, current_user)
    require_staff_permission(current_user, PERMISSION_REPORT)

    # Đối soát là việc giải quyết tiền đã phát sinh nên luôn được nhìn toàn bộ,
    # kể cả sau khi Pro hết hạn. Dashboard bình thường của Free chỉ xem 31 ngày.
    if not reconciliation_only:
        tu_ngay, den_ngay = _limit_free_report_range(
            db, shop_id, tu_ngay, den_ngay
        )

    if page < 1:
        raise HTTPException(status_code=400, detail=tr("page phải >= 1"))
    if per_page < 1 or per_page > MAX_PAGE_SIZE:
        raise HTTPException(
            status_code=400,
            detail=tr("per_page phải từ 1 đến {maximum}", maximum=MAX_PAGE_SIZE),
        )

    open_reconciliation = or_(
        models.Order.status == order_service.STATUS_UNRECONCILED,
        and_(
            models.Order.refund_due_amount > order_service.MONEY_EPSILON,
            models.Order.refund_completed_at.is_(None),
        ),
        # F6: tiền đã về cho một đơn ghi nợ nhưng webhook không tự áp vào đơn
        # (bẫy 25). Không đưa lên đây thì khoản đó chỉ nằm trong `SystemLog` -
        # tiền về mà người bán không biết để đi thu.
        models.Order.id.in_(_don_co_tien_ve_chua_ghi_nhan(db, shop_id)),
    )
    reconciliation_count = (
        db.query(models.Order)
        .filter(models.Order.shop_id == shop_id, open_reconciliation)
        .count()
    )

    base = db.query(models.Order).filter(models.Order.shop_id == shop_id)
    base = _loc_khoang_ngay(base, tu_ngay, den_ngay)
    if reconciliation_only:
        base = base.filter(open_reconciliation)

    tong_don = base.count()
    orders = (
        base.options(joinedload(models.Order.created_by))
        .order_by(models.Order.created_at.desc())
        .offset((page - 1) * per_page)
        .limit(per_page)
        .all()
    )

    if reconciliation_only:
        # Chế độ Đối Soát được phép nhìn các giao dịch cũ đang cần xử lý, nhưng
        # không vì thế mà biến thành cửa hậu đọc tổng doanh thu toàn lịch sử của
        # gói Pro. Chỉ cộng đúng các dòng PAID đang nằm trong tập đối soát mở.
        doanh_thu = (
            base.with_entities(func.sum(models.Order.total_amount))
            .filter(models.Order.status == "PAID")
            .scalar()
            or 0
        )
    else:
        doanh_thu = (
            _loc_khoang_ngay(
                db.query(func.sum(models.Order.total_amount)).filter(
                    models.Order.shop_id == shop_id, models.Order.status == "PAID"
                ),
                tu_ngay,
                den_ngay,
            ).scalar()
            or 0
        )

    # Chỉ hỏi tiền-về-chưa-ghi-nhận cho ĐÚNG các đơn của trang này, không quét
    # cả shop: màn Đối Soát phân trang, mà số đơn thì lớn dần theo thời gian.
    tien_ve = _tien_ve_chua_ghi_nhan_theo_don(db, [o.id for o in orders])

    return {
        "total_revenue": doanh_thu,
        "orders": [_dashboard_order(o, tien_ve.get(o.id)) for o in orders],
        "page": page,
        "per_page": per_page,
        "total_orders": tong_don,
        "has_more": page * per_page < tong_don,
        "reconciliation_count": reconciliation_count,
    }


def _tien_ve_chua_ghi_nhan_theo_don(
    db: Session, order_ids: List[int]
) -> Dict[int, float]:
    """Tổng tiền `BANK_UNAPPLIED` của từng đơn trong danh sách."""
    if not order_ids:
        return {}
    hang = (
        db.query(
            models.OrderPayment.order_id,
            func.sum(models.OrderPayment.amount),
        )
        .filter(
            models.OrderPayment.order_id.in_(order_ids),
            models.OrderPayment.entry_type
            == order_service.ENTRY_BANK_UNAPPLIED,
        )
        .group_by(models.OrderPayment.order_id)
        .all()
    )
    return {order_id: float(tong or 0) for order_id, tong in hang}


def _dashboard_order(
    order: models.Order, tien_ve_chua_ghi_nhan: Optional[float] = None
) -> Dict[str, Any]:
    result = {
        "id": order.id,
        "total": order.total_amount,
        "status": order.status,
        "date": order.created_at,
        "cashier_username": order.created_by.username if order.created_by else None,
        "shift_id": order.shift_id,
        # F6: tiền đã về tài khoản cho đơn này nhưng webhook KHÔNG áp vào đơn.
        # 0 nghĩa là không có, chứ không phải "chưa biết" - đây là tổng của một
        # tập bút toán đếm được.
        "unapplied_transfer_amount": float(tien_ve_chua_ghi_nhan or 0),
    }
    result.update(order_service.payment_summary(order))
    return result


def admin_dashboard(db: Session) -> List[Dict[str, Any]]:
    shops = db.query(models.Shop).all()
    log_to_file(f"get_admin_dashboard: Found {len(shops)} shops in DB")
    return [
        {"shop_name": s.name, "total_revenue": _paid_revenue(db, s.id)} for s in shops
    ]


def shop_stats(
    db: Session,
    current_user: models.User,
    shop_id: int,
    tu_ngay: Optional[str] = None,
    den_ngay: Optional[str] = None,
) -> Dict[str, Any]:
    """Thống kê của shop. Không truyền ngày -> toàn bộ lịch sử + xu hướng 7 ngày
    (đúng như trước). Truyền ngày -> mọi con số và biểu đồ đều theo khoảng đó."""
    shop = require_shop_access(db, shop_id, current_user)
    require_staff_permission(current_user, PERMISSION_REPORT)
    tu_ngay, den_ngay = _limit_free_report_range(
        db, shop_id, tu_ngay, den_ngay
    )
    co_loc_ngay = bool((tu_ngay or "").strip() or (den_ngay or "").strip())

    total_rev = (
        _loc_khoang_ngay(
            db.query(func.sum(models.Order.total_amount)).filter(
                models.Order.shop_id == shop_id, models.Order.status == "PAID"
            ),
            tu_ngay,
            den_ngay,
        ).scalar()
        or 0
    )
    total_orders = _loc_khoang_ngay(
        db.query(models.Order).filter(models.Order.shop_id == shop_id), tu_ngay, den_ngay
    ).count()

    paid_orders_subquery = _loc_khoang_ngay(
        db.query(models.Order.id).filter(
            models.Order.shop_id == shop_id, models.Order.status == "PAID"
        ),
        tu_ngay,
        den_ngay,
    )
    total_sold = (
        db.query(func.sum(models.OrderItem.quantity))
        .filter(models.OrderItem.order_id.in_(paid_orders_subquery))
        .scalar()
        or 0
    )

    # Gộp các biến thể của cùng một nhóm thành MỘT dòng. Không gộp thì một cái
    # áo có 4 size chiếm 4 trong 5 chỗ của bảng "bán chạy nhất", đẩy hết mặt
    # hàng khác ra ngoài - càng nhiều biến thể bảng càng vô dụng, mà biến thể
    # chính là thứ vừa được thêm ở F6.
    #
    # Hàng đơn lẻ vẫn gom theo `product_name` ĐÃ CHỐT LÚC BÁN như trước, không
    # đổi. Chỉ hàng có `variant_group` mới gom theo nhóm HIỆN TẠI của sản phẩm -
    # vì "áo thun bán được bao nhiêu" là câu hỏi về danh mục hôm nay, không phải
    # về cái tên hồi tháng trước. Dòng đơn hàng cũ không có `product_id`
    # (trước migration A1a) rơi về `product_name` nhờ COALESCE.
    _ten_gom = func.coalesce(
        models.Product.variant_group, models.OrderItem.product_name
    )
    top_products_query = (
        db.query(
            _ten_gom.label("ten"),
            func.sum(models.OrderItem.quantity).label("so_luong"),
            func.count(distinct(models.OrderItem.product_id)).label("so_bien_the"),
            # NULL = mọi dòng trong cụm này đều là hàng đơn lẻ.
            func.max(models.Product.variant_group).label("nhom"),
        )
        .outerjoin(models.Product, models.Product.id == models.OrderItem.product_id)
        .filter(models.OrderItem.order_id.in_(paid_orders_subquery))
        .group_by(_ten_gom)
        .order_by(func.sum(models.OrderItem.quantity).desc())
        .limit(5)
        .all()
    )
    top_products = [
        {
            "name": r.ten,
            "qty": r.so_luong,
            # 0 = hàng đơn lẻ. Giao diện chỉ ghi "(N loại)" khi > 0, nếu không
            # mọi món trong tiệm đều bị dán thêm "(1 loại)" vô nghĩa.
            "variants": r.so_bien_the if r.nhom else 0,
        }
        for r in top_products_query
    ]

    if co_loc_ngay:
        # Biểu đồ chạy theo đúng khoảng người dùng chọn
        recent_orders = _loc_khoang_ngay(
            db.query(models.Order).filter(
                models.Order.shop_id == shop_id, models.Order.status == "PAID"
            ),
            tu_ngay,
            den_ngay,
        ).all()
        revenue_by_date = {o.created_at.strftime("%Y-%m-%d"): 0 for o in recent_orders}
    else:
        # Mặc định: 7 ngày gần nhất (giữ nguyên hành vi cũ)
        seven_days_ago = datetime.utcnow() - timedelta(days=TREND_DAYS - 1)
        recent_orders = (
            db.query(models.Order)
            .filter(
                models.Order.shop_id == shop_id,
                models.Order.status == "PAID",
                models.Order.created_at >= seven_days_ago,
            )
            .all()
        )
        revenue_by_date = {
            (datetime.utcnow() - timedelta(days=i)).strftime("%Y-%m-%d"): 0
            for i in range(TREND_DAYS)
        }

    for o in recent_orders:
        d_str = o.created_at.strftime("%Y-%m-%d")
        if d_str in revenue_by_date:
            revenue_by_date[d_str] += o.total_amount

    trend_labels = sorted(revenue_by_date.keys())
    trend_data = [revenue_by_date[k] for k in trend_labels]

    tra_hang = _tra_hang_anh_huong_lai(db, shop_id, tu_ngay, den_ngay)
    ket_qua = {
        "total_revenue": total_rev,
        "total_orders": total_orders,
        "total_sold": total_sold,
        "top_products": top_products,
        "trend_labels": trend_labels,
        "trend_data": trend_data,
        # `total_revenue` giữ nguyên nghĩa cũ là tiền bán ra trong kỳ; phần
        # khách trả lại đứng riêng để không âm thầm đổi nghĩa một con số mà
        # người dùng đã quen đọc.
        "returned_amount": tra_hang["returned_amount"],
        "net_revenue": total_rev - tra_hang["returned_amount"],
        # F4: tổng tiền khách còn nợ. CỐ Ý đứng riêng, KHÔNG cộng vào doanh thu:
        # doanh thu ở đây là tiền đã thực thu, còn đây là tiền mới hứa trả.
        # Cũng CỐ Ý không lọc theo khoảng ngày - nợ là số dư tại thời điểm hiện
        # tại, không phải phát sinh trong kỳ.
        "receivable_amount": _cong_no_phai_thu(db, shop_id),
    }
    # MANAGER có PERMISSION_REPORT nên vẫn xem được doanh thu, nhưng lãi thì
    # không: biết lãi là suy ra được giá vốn. Khi không có quyền thì BỎ HẲN các
    # field này khỏi phản hồi, không trả 0 - ở đây 0 là một con số có nghĩa
    # (bán đúng bằng giá vốn), trả 0 là nói dối chứ không phải giấu.
    if has_cost_visibility(shop, current_user):
        # Hủy hàng chỉ hiện cho người xem được giá vốn: số lỗ chính là giá vốn
        # nhân số lượng, nên nói ra nó là nói ra giá vốn.
        ket_qua.update(
            _lai_gop_da_dieu_chinh(
                db, shop_id, paid_orders_subquery, tu_ngay, den_ngay, tra_hang
            )
        )
    return ket_qua


def _lai_gop_da_dieu_chinh(
    db: Session,
    shop_id: int,
    paid_orders_subquery,
    tu_ngay: Optional[str],
    den_ngay: Optional[str],
    tra_hang: Dict[str, Any],
) -> Dict[str, Any]:
    """Lãi gộp đã trừ hàng trả lại và hàng hủy.

    Dùng chung cho Thống Kê và Dòng Tiền. Hai màn nói hai con số khác nhau về
    cùng một tháng là lúc người ta thôi tin cả hai, nên phép tính chỉ được viết
    ở đúng một chỗ.
    """
    huy_hang = _huy_hang_anh_huong_lai(db, shop_id, tu_ngay, den_ngay)
    lai = _lai_gop(db, paid_orders_subquery)
    lai["gross_profit"] -= tra_hang["profit_reduction"]
    lai["gross_profit"] -= huy_hang["write_off_loss"]
    lai["returns_missing_cost"] = tra_hang["returns_missing_cost"]
    lai.update(huy_hang)
    # Tỷ suất tính lại trên doanh thu ĐÃ TRỪ hàng trả: giữ tử số mới mà mẫu
    # số cũ sẽ ra một con số không nói lên điều gì.
    mau_so = lai["revenue_with_cost"] - tra_hang["returned_amount"]
    lai["gross_margin"] = (
        (lai["gross_profit"] / mau_so * 100)
        if mau_so > order_service.MONEY_EPSILON
        else None
    )
    return lai


# ---------------------------------------------------------------------------
# K1: Lợi nhuận ròng và dòng tiền thực
# ---------------------------------------------------------------------------
#
# HAI CON SỐ KHÁC NHAU, tuyệt đối không gộp:
#
#   Lợi nhuận ròng = lãi gộp (đã trừ hàng trả, hàng hủy) - chi phí vận hành
#   Dòng tiền thực = mọi đồng vào - mọi đồng ra, trong kỳ
#
# Nhập hàng 10 triệu trả tiền ngay mà chưa bán món nào: dòng tiền -10 triệu,
# lãi ròng không đổi (hàng còn trong kho, chỉ thành giá vốn lúc bán). Bán 5
# triệu ghi nợ: lãi tăng, dòng tiền đứng yên. Gộp hai con số này là nói dối
# theo cả hai hướng tùy tháng.

# Tiền VÀO. `BANK_UNAPPLIED` CỐ Ý không có mặt: đó là tiền về tài khoản cho một
# đơn ghi nợ mà webhook không áp vào đơn - nó không phải khoản thu (xem
# order_service), và khi người bán ghi nhận thu nợ thì `DEBT_*` mới là bút toán
# thật. Cộng cả hai là đếm một lần chuyển khoản thành hai lần tiền vào.
_CASHFLOW_IN_GROUPS = (
    ("sale_cash", "Bán hàng thu tiền mặt", (order_service.ENTRY_CASH,)),
    ("sale_bank", "Bán hàng chuyển khoản", (order_service.ENTRY_BANK,)),
    (
        "debt_collected",
        "Khách trả nợ",
        (order_service.ENTRY_DEBT_CASH, order_service.ENTRY_DEBT_TRANSFER),
    ),
)

# Tiền RA từ sổ đơn hàng: hoàn khoản chuyển thừa và hoàn tiền hàng khách trả.
_CASHFLOW_OUT_ORDER_ENTRIES = (
    order_service.ENTRY_REFUND_CASH,
    order_service.ENTRY_REFUND_TRANSFER,
    return_service.ENTRY_RETURN_CASH,
    return_service.ENTRY_RETURN_TRANSFER,
)


def _ngay_vn(cot):
    """Cột datetime UTC -> ngày lịch Việt Nam, để gom nhóm biểu đồ theo ngày.

    Việt Nam là UTC+7 quanh năm, không có giờ mùa hè, nên phép cộng cố định này
    đúng cho mọi ngày mà dữ liệu của app có thể có.
    """
    return func.date(cot, "+7 hours")


def _cong_don_theo_ngay(dich: Dict[str, float], hang) -> float:
    """Cộng kết quả (ngày, số tiền) vào một dict, trả về tổng vừa cộng."""
    tong = 0.0
    for ngay, tien in hang:
        if not ngay:
            continue
        so = float(tien or 0)
        dich[ngay] = dich.get(ngay, 0.0) + so
        tong += so
    return tong


def _cash_movement_da_thuoc_chung_tu(db: Session, shop_id: int):
    """Các chuyển động két đã được đại diện bởi một chứng từ khác.

    Trả nhà cung cấp và chi phí vận hành bằng tiền mặt đều sinh đúng một
    `CashMovement` hướng OUT. Nếu dòng tiền cộng cả chứng từ lẫn chuyển động
    thì mỗi lần chi bị đếm hai lần. Đây là danh sách để loại ra, giữ lại đúng
    các khoản thu/chi tay không thuộc chứng từ nào.

    Trả về hai `Query` để caller đưa thẳng vào `in_()`; KHÔNG gọi `.subquery()`
    (bẫy 12).
    """
    tu_ncc = db.query(models.SupplierPayment.cash_movement_id).filter(
        models.SupplierPayment.shop_id == shop_id,
        models.SupplierPayment.cash_movement_id.isnot(None),
    )
    tu_chi_phi = db.query(models.OperatingExpense.cash_movement_id).filter(
        models.OperatingExpense.shop_id == shop_id,
        models.OperatingExpense.cash_movement_id.isnot(None),
    )
    return tu_ncc, tu_chi_phi


def _nhan_bieu_do(
    co_du_lieu: List[str],
    tu_ngay: Optional[str],
    den_ngay: Optional[str],
) -> List[str]:
    """Trục ngày của biểu đồ: liền mạch, không nhảy cóc qua ngày không bán.

    Ngày nghỉ mà biến mất khỏi trục thì đường cộng dồn trông như đi ngang trong
    khi thực tế là không có gì xảy ra - hai chuyện khác nhau. Kỳ quá dài thì
    quay về chỉ vẽ ngày có phát sinh, vì vài trăm cột đứng cạnh nhau cũng không
    đọc được gì.
    """
    moc_dau = _parse_ngay(tu_ngay, "tu_ngay")
    moc_cuoi = _parse_ngay(den_ngay, "den_ngay")
    if co_du_lieu:
        thap_nhat = datetime.strptime(min(co_du_lieu), "%Y-%m-%d").date()
        cao_nhat = datetime.strptime(max(co_du_lieu), "%Y-%m-%d").date()
        moc_dau = min(moc_dau, thap_nhat) if moc_dau else thap_nhat
        moc_cuoi = max(moc_cuoi, cao_nhat) if moc_cuoi else cao_nhat
    if moc_dau is None or moc_cuoi is None or moc_dau > moc_cuoi:
        return sorted(co_du_lieu)
    so_ngay = (moc_cuoi - moc_dau).days + 1
    if so_ngay > 400:
        return sorted(co_du_lieu)
    return [
        (moc_dau + timedelta(days=i)).isoformat() for i in range(so_ngay)
    ]


def _dong_tien(
    db: Session,
    shop_id: int,
    tu_ngay: Optional[str],
    den_ngay: Optional[str],
) -> Dict[str, Any]:
    """Mọi đồng vào và ra của shop trong kỳ, kèm chuỗi số liệu theo ngày.

    CỐ Ý không có tiền gói Pro ở đây (bẫy 33): đó là tiền nền tảng, đi qua
    `SubscriptionPayment` và tài khoản riêng, không phải tiền của cửa hàng.
    """
    vao: Dict[str, Dict[str, float]] = {}
    ra: Dict[str, Dict[str, float]] = {}
    tong_vao: Dict[str, float] = {}
    tong_ra: Dict[str, float] = {}

    don_cua_shop = db.query(models.Order.id).filter(
        models.Order.shop_id == shop_id
    )

    for khoa, _nhan, entries in _CASHFLOW_IN_GROUPS:
        theo_ngay: Dict[str, float] = {}
        hang = _loc_khoang_ngay(
            db.query(
                _ngay_vn(models.OrderPayment.created_at),
                func.sum(models.OrderPayment.amount),
            ).filter(
                models.OrderPayment.order_id.in_(don_cua_shop),
                models.OrderPayment.entry_type.in_(entries),
            ),
            tu_ngay,
            den_ngay,
            cot=models.OrderPayment.created_at,
        ).group_by(_ngay_vn(models.OrderPayment.created_at)).all()
        tong_vao[khoa] = _cong_don_theo_ngay(theo_ngay, hang)
        vao[khoa] = theo_ngay

    # Hoàn tiền khách (chuyển thừa + trả hàng).
    hoan: Dict[str, float] = {}
    tong_ra["refund"] = _cong_don_theo_ngay(
        hoan,
        _loc_khoang_ngay(
            db.query(
                _ngay_vn(models.OrderPayment.created_at),
                func.sum(models.OrderPayment.amount),
            ).filter(
                models.OrderPayment.order_id.in_(don_cua_shop),
                models.OrderPayment.entry_type.in_(_CASHFLOW_OUT_ORDER_ENTRIES),
            ),
            tu_ngay,
            den_ngay,
            cot=models.OrderPayment.created_at,
        ).group_by(_ngay_vn(models.OrderPayment.created_at)).all(),
    )
    ra["refund"] = hoan

    # Trả nhà cung cấp. Đếm CHỨNG TỪ chứ không đếm chuyển động két, và mọi
    # phương thức đều là tiền ra khỏi túi chủ shop (kể cả `OUTSIDE`).
    ncc: Dict[str, float] = {}
    tong_ra["supplier"] = _cong_don_theo_ngay(
        ncc,
        _loc_khoang_ngay(
            db.query(
                _ngay_vn(models.SupplierPayment.created_at),
                func.sum(models.SupplierPayment.amount),
            ).filter(models.SupplierPayment.shop_id == shop_id),
            tu_ngay,
            den_ngay,
            cot=models.SupplierPayment.created_at,
        ).group_by(_ngay_vn(models.SupplierPayment.created_at)).all(),
    )
    ra["supplier"] = ncc

    # Chi phí vận hành: lấy nguyên số đã trả vào đúng NGÀY CHI, không phân bổ.
    # Trả trước 30 triệu tiền nhà thì dòng tiền phải thấy đủ 30 triệu ra hôm đó.
    chi_phi = expense_service.tien_chi_theo_ngay(db, shop_id, tu_ngay, den_ngay)
    ra["expense"] = {k: float(v) for k, v in chi_phi["by_date"].items()}
    tong_ra["expense"] = float(chi_phi["total"])

    # Thu/chi tay trong ca, sau khi loại các dòng đã thuộc chứng từ ở trên.
    tu_ncc, tu_chi_phi = _cash_movement_da_thuoc_chung_tu(db, shop_id)
    ca_cua_shop = db.query(models.CashShift.id).filter(
        models.CashShift.shop_id == shop_id
    )
    for khoa, huong in (
        ("cash_topup", shift_service.DIRECTION_IN),
        ("cash_withdraw", shift_service.DIRECTION_OUT),
    ):
        theo_ngay: Dict[str, float] = {}
        tong = _cong_don_theo_ngay(
            theo_ngay,
            _loc_khoang_ngay(
                db.query(
                    _ngay_vn(models.CashMovement.created_at),
                    func.sum(models.CashMovement.amount),
                ).filter(
                    models.CashMovement.shift_id.in_(ca_cua_shop),
                    models.CashMovement.direction == huong,
                    ~models.CashMovement.id.in_(tu_ncc),
                    ~models.CashMovement.id.in_(tu_chi_phi),
                ),
                tu_ngay,
                den_ngay,
                cot=models.CashMovement.created_at,
            ).group_by(_ngay_vn(models.CashMovement.created_at)).all(),
        )
        if huong == shift_service.DIRECTION_IN:
            vao[khoa] = theo_ngay
            tong_vao[khoa] = tong
        else:
            ra[khoa] = theo_ngay
            tong_ra[khoa] = tong

    nhan_vao = {
        "sale_cash": "Bán hàng thu tiền mặt",
        "sale_bank": "Bán hàng chuyển khoản",
        "debt_collected": "Khách trả nợ",
        "cash_topup": "Bỏ thêm tiền vào két",
    }
    nhan_ra = {
        "supplier": "Trả tiền nhập hàng",
        "expense": "Chi phí vận hành",
        "refund": "Hoàn tiền cho khách",
        "cash_withdraw": "Rút tiền khỏi két",
    }

    moi_ngay = sorted(
        {ngay for nhom in list(vao.values()) + list(ra.values()) for ngay in nhom}
    )
    nhan = _nhan_bieu_do(moi_ngay, tu_ngay, den_ngay)

    chuoi_vao = [
        round(sum(nhom.get(ngay, 0.0) for nhom in vao.values()), 2)
        for ngay in nhan
    ]
    chuoi_ra = [
        round(sum(nhom.get(ngay, 0.0) for nhom in ra.values()), 2)
        for ngay in nhan
    ]
    cong_don: List[float] = []
    chay = 0.0
    for i in range(len(nhan)):
        chay += chuoi_vao[i] - chuoi_ra[i]
        cong_don.append(round(chay, 2))

    return {
        "cash_in_total": sum(tong_vao.values()),
        "cash_out_total": sum(tong_ra.values()),
        "net_cashflow": sum(tong_vao.values()) - sum(tong_ra.values()),
        "cash_in_breakdown": [
            {"key": k, "label": tr(nhan_vao[k]), "amount": v}
            for k, v in sorted(tong_vao.items(), key=lambda r: -r[1])
            if v > order_service.MONEY_EPSILON
        ],
        "cash_out_breakdown": [
            {"key": k, "label": tr(nhan_ra[k]), "amount": v}
            for k, v in sorted(tong_ra.items(), key=lambda r: -r[1])
            if v > order_service.MONEY_EPSILON
        ],
        "chart": {
            "labels": nhan,
            "cash_in": chuoi_vao,
            "cash_out": chuoi_ra,
            # Cộng dồn TRONG KỲ, bắt đầu từ 0 - không phải số dư tuyệt đối của
            # két. Giao diện phải ghi rõ điều đó, nếu không người xem sẽ đọc
            # đường này thành "tiền tôi đang có".
            "cumulative": cong_don,
        },
        "supplier_payment_total": tong_ra.get("supplier", 0.0),
    }


def net_cashflow_report(
    db: Session,
    current_user: models.User,
    shop_id: int,
    tu_ngay: Optional[str] = None,
    den_ngay: Optional[str] = None,
) -> Dict[str, Any]:
    """Màn Dòng Tiền: lợi nhuận ròng, dòng tiền thực và lý do chúng khác nhau.

    CHỈ chủ shop và ADMIN. Biết chi phí và lãi ròng là suy ngược ra được giá
    vốn, nên ranh giới quyền ở đây phải trùng với `require_cost_visibility` chứ
    không nới theo PERMISSION_REPORT như Thống Kê.
    """
    shop = require_shop_access(db, shop_id, current_user)
    require_cost_visibility(shop, current_user)
    # Gói Free vẫn nhập chi phí và xem 31 ngày gần nhất bình thường; chỉ kỳ cũ
    # hơn mới cần Pro, đúng cùng chính sách với báo cáo hiện có.
    tu_ngay, den_ngay = _limit_free_report_range(db, shop_id, tu_ngay, den_ngay)

    doanh_thu = (
        _loc_khoang_ngay(
            db.query(func.sum(models.Order.total_amount)).filter(
                models.Order.shop_id == shop_id, models.Order.status == "PAID"
            ),
            tu_ngay,
            den_ngay,
        ).scalar()
        or 0
    )
    paid_orders_subquery = _loc_khoang_ngay(
        db.query(models.Order.id).filter(
            models.Order.shop_id == shop_id, models.Order.status == "PAID"
        ),
        tu_ngay,
        den_ngay,
    )
    tra_hang = _tra_hang_anh_huong_lai(db, shop_id, tu_ngay, den_ngay)
    lai = _lai_gop_da_dieu_chinh(
        db, shop_id, paid_orders_subquery, tu_ngay, den_ngay, tra_hang
    )
    chi_phi = expense_service.tong_hop_chi_phi(db, shop_id, tu_ngay, den_ngay)
    lai_rong = lai["gross_profit"] - chi_phi["operating_expense_total"]

    tien = _dong_tien(db, shop_id, tu_ngay, den_ngay)
    cong_no = _cong_no_phai_thu(db, shop_id)

    # Vì sao lãi ròng khác dòng tiền. CỐ Ý không phải một phép cân đối khép kín
    # (số dư đầu kỳ, hàng tồn, công nợ nhiều kỳ đan vào nhau, cân cho khớp tuyệt
    # đối sẽ ra một con số "chênh lệch khác" mà không ai giải thích nổi). Đây là
    # danh sách các NGUYÊN NHÂN LỚN, mỗi con số đều tra ngược về chứng từ được,
    # và giao diện chỉ đọc ra những dòng khác 0.
    ly_do = [
        {
            "key": "purchase_paid",
            "amount": tien["supplier_payment_total"],
            "label": tr(
                "tiền nhập hàng đã trả, hàng còn trong kho nên chưa tính lời lỗ"
            ),
        },
        {
            "key": "prepaid",
            "amount": float(chi_phi["prepaid_remaining"]),
            "label": tr(
                "tiền đã trả trước cho các tháng sau, chưa tính vào lãi tháng này"
            ),
        },
        {
            "key": "receivable",
            "amount": float(cong_no),
            "label": tr("khách còn nợ, đã bán nhưng chưa cầm được tiền"),
        },
    ]

    ket_qua = {
        "from_date": tu_ngay,
        "to_date": den_ngay,
        "total_revenue": doanh_thu,
        "returned_amount": tra_hang["returned_amount"],
        "net_revenue": doanh_thu - tra_hang["returned_amount"],
        "receivable_amount": cong_no,
        "net_profit": lai_rong,
        "difference_notes": [
            r for r in ly_do if r["amount"] > order_service.MONEY_EPSILON
        ],
    }
    ket_qua.update(lai)
    ket_qua.update(chi_phi)
    ket_qua.update(tien)
    return ket_qua


def _workbook_to_stream(wb: "openpyxl.Workbook") -> io.BytesIO:
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def admin_excel(db: Session) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = tr("Doanh thu Shops")
    ws.append([tr("Tên Shop"), tr("Tổng Doanh Thu")])
    for s in db.query(models.Shop).all():
        ws.append([s.name, _paid_revenue(db, s.id)])
    return _workbook_to_stream(wb)


def _gia_von_don(order: models.Order) -> Optional[float]:
    """Tổng giá vốn của một đơn, hoặc None nếu còn dòng chưa khai giá vốn.

    Thiếu một dòng là cả đơn không tính được: cộng phần đã biết rồi so với
    doanh thu cả đơn sẽ ra một con số lãi cao hơn sự thật.
    """
    if not order.items:
        return None
    tong = 0.0
    for item in order.items:
        if item.cost_price is None:
            return None
        tong += item.cost_price * item.quantity
    return tong


def seller_excel(db: Session, current_user: models.User, shop_id: int) -> io.BytesIO:
    shop = require_shop_access(db, shop_id, current_user)
    require_staff_permission(current_user, PERMISSION_REPORT)
    subscription_service.require_pro(db, shop_id)
    xem_duoc_gia_von = has_cost_visibility(shop, current_user)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = tr("Lịch sử giao dịch")
    tieu_de = [
        tr("Mã đơn"),
        tr("Ngày tạo"),
        tr("Thu ngân"),
        tr("Mã ca"),
        tr("Trạng thái"),
        tr("Thành tiền"),
    ]
    if xem_duoc_gia_von:
        tieu_de += [tr("Giá vốn"), tr("Lãi gộp")]
    ws.append(tieu_de)

    orders = (
        db.query(models.Order)
        .options(
            joinedload(models.Order.created_by),
            joinedload(models.Order.items),
        )
        .filter(models.Order.shop_id == shop_id)
        .all()
    )
    total_rev = 0
    tong_gia_von = 0.0
    tong_lai = 0.0
    don_thieu_gia_von = 0
    for o in orders:
        dong = [
            o.id,
            str(o.created_at),
            o.created_by.username if o.created_by else "",
            o.shift_id or "",
            tr({
                "PENDING": "Chờ thanh toán",
                "PAID": "Đã thanh toán",
                "CANCELLED": "Đã hủy",
                "UNRECONCILED": "Cần đối soát",
            }.get(o.status, o.status)),
            o.total_amount,
        ]
        gia_von = _gia_von_don(o) if xem_duoc_gia_von else None
        if xem_duoc_gia_von:
            if gia_von is None:
                # Ô trống hơn hẳn số 0: 0 trong cột giá vốn đọc ra là hàng tặng.
                dong += ["", tr("Chưa khai giá vốn")]
            else:
                dong += [gia_von, o.total_amount - gia_von]
        ws.append(dong)
        if o.status == "PAID":
            total_rev += o.total_amount
            if gia_von is None:
                if xem_duoc_gia_von:
                    don_thieu_gia_von += 1
            else:
                tong_gia_von += gia_von
                tong_lai += o.total_amount - gia_von

    ws.append([])
    ws.append([tr("Tổng Doanh Thu (Đã thanh toán)"), total_rev])
    if xem_duoc_gia_von:
        ws.append([tr("Tổng Giá Vốn (đơn đã đủ giá vốn)"), tong_gia_von])
        ws.append([tr("Tổng Lãi Gộp (đơn đã đủ giá vốn)"), tong_lai])
        if don_thieu_gia_von:
            ws.append([
                tr("Số đơn chưa đủ giá vốn (không tính vào lãi)"),
                don_thieu_gia_von,
            ])
        _them_sheet_hang_huy(db, wb, shop_id)
    return _workbook_to_stream(wb)


def _them_sheet_hang_huy(db: Session, wb, shop_id: int) -> None:
    """Sheet thứ hai: hàng ra khỏi kho mà KHÔNG bán.

    Sheet đầu chỉ có đơn hàng, nên đọc một mình nó thì tồn kho giảm mà không
    hiểu vì sao, và số lãi gộp ở cuối sheet đó cao hơn con số trên Dashboard
    đúng bằng phần vốn đã hủy. Hai chỗ nói hai số khác nhau về cùng một tháng
    là lúc người ta thôi tin cả hai.

    CHỈ thêm khi người xuất được xem giá vốn - số lỗ chính là giá vốn nhân số
    lượng. Caller đã kiểm, ở đây không kiểm lại nhưng cũng đừng gọi chỗ khác.

    Một dòng cho MỘT LÔ bị hủy, không phải một dòng cho một phiếu: đối chiếu
    cuối tháng cần biết từng hạn, và Excel thì lọc theo cột dễ hơn đọc ô gộp.
    """
    phieu = (
        db.query(models.StockWriteOff)
        .filter(models.StockWriteOff.shop_id == shop_id)
        .order_by(models.StockWriteOff.id.desc())
        .all()
    )
    ws = wb.create_sheet(tr("Hàng đã hủy"))
    ws.append([
        tr("Mã phiếu"),
        tr("Ngày hủy"),
        tr("Lý do"),
        tr("Sản phẩm"),
        tr("Hạn sử dụng"),
        tr("Số lượng"),
        tr("Giá vốn đơn vị"),
        tr("Thành tiền"),
        tr("Người bấm"),
        tr("Ghi chú"),
    ])
    if not phieu:
        return

    dong_theo_phieu: Dict[int, List[models.StockWriteOffItem]] = {}
    for d in (
        db.query(models.StockWriteOffItem)
        .filter(
            models.StockWriteOffItem.write_off_id.in_([p.id for p in phieu])
        )
        .order_by(models.StockWriteOffItem.id)
        .all()
    ):
        dong_theo_phieu.setdefault(d.write_off_id, []).append(d)

    ten_nguoi = {
        uid: ten
        for uid, ten in db.query(models.User.id, models.User.username).filter(
            models.User.id.in_(
                [p.created_by_user_id for p in phieu if p.created_by_user_id]
            )
        )
    } if any(p.created_by_user_id for p in phieu) else {}

    nhan_ly_do = {
        write_off_service.REASON_EXPIRED: "Hết hạn",
        write_off_service.REASON_DAMAGED: "Hỏng / vỡ",
        write_off_service.REASON_LOST: "Thất thoát",
    }

    tong_so_luong = 0
    tong_lo = 0.0
    phieu_thieu_gia_von = 0
    for p in phieu:
        dong = dong_theo_phieu.get(p.id, [])
        thieu = any(d.cost_price is None for d in dong)
        if thieu:
            phieu_thieu_gia_von += 1
        for d in dong:
            # Ô TRỐNG chứ không phải 0 khi chưa khai giá vốn: 0 trong cột tiền
            # đọc ra là "hàng này không đáng đồng nào", khác hẳn "chưa ai khai".
            co_gia = d.cost_price is not None
            ws.append([
                p.id,
                str(p.created_at),
                tr(nhan_ly_do.get(p.reason, p.reason)),
                d.product_name or "",
                d.expiry_date or "",
                d.quantity,
                d.cost_price if co_gia else "",
                (float(d.cost_price) * int(d.quantity or 0)) if co_gia else
                tr("Chưa khai giá vốn"),
                ten_nguoi.get(p.created_by_user_id, ""),
                p.note or "",
            ])
        tong_so_luong += int(p.total_quantity or 0)
        if not thieu:
            tong_lo += sum(
                float(d.cost_price) * int(d.quantity or 0) for d in dong
            )

    ws.append([])
    ws.append([tr("Tổng số lượng đã hủy"), tong_so_luong])
    ws.append([tr("Tổng lỗ (phiếu đã đủ giá vốn)"), tong_lo])
    if phieu_thieu_gia_von:
        ws.append([
            tr("Số phiếu chưa đủ giá vốn (không tính vào lỗ)"),
            phieu_thieu_gia_von,
        ])
